"""Converts the data pack's .xlsx and decision files into one JSON cache.

Run once from src/backend/:  python scripts/build-data-cache.py
Writes src/backend/data/enrichment.json (gitignored).
"""
import json
import pathlib
import openpyxl

ROOT = pathlib.Path(__file__).resolve().parents[3]
PACK = ROOT / "zollhof-recognition-data-pack"
OUT = pathlib.Path(__file__).resolve().parents[1] / "data" / "enrichment.json"

# 1. KldB 2010 (5-digit) -> ISCO-08 (4-digit). Sheet index 1, header row 5, data from row 6.
#    Inverted to ISCO-4 -> set of KldB-4 prefixes, which is how we query it.
isco_to_kldb: dict[str, set[str]] = {}
wb = openpyxl.load_workbook(PACK / "crosswalks" / "kldb2010-isco08-crosswalk.xlsx", read_only=True)
for row in wb[wb.sheetnames[1]].iter_rows(min_row=6, values_only=True):
    kldb5, isco4 = row[0], row[3]
    if not kldb5 or not isco4:
        continue
    isco_to_kldb.setdefault(str(isco4).strip(), set()).add(str(kldb5).strip()[:4])

# 2. BA shortage analysis. Three sheets (Fachkraefte/Spezialisten/Experten), data from row 11.
#    Column A is "<KldB-4> <name>", column T (index 19) is "Durchschnittliche Punktezahl".
#    BA's methodology treats an average score >= 2.0 as an Engpassberuf.
kldb_score: dict[str, float] = {}
wb = openpyxl.load_workbook(
    PACK / "labour-market" / "ba" / "2025_Deutschland_Engpass.xlsx", read_only=True, data_only=True
)
for name in wb.sheetnames:
    for row in wb[name].iter_rows(min_row=11, values_only=True):
        code = str(row[0])[:4] if row[0] else ""
        score = row[19] if len(row) > 19 else None
        if code.isdigit() and isinstance(score, (int, float)):
            kldb_score[code] = max(kldb_score.get(code, 0.0), float(score))

# 3. Regulated professions, keyed by ISCO-4.
#    regprof-german-professions.json has no ISCO codes, so we use the decisions
#    file instead: it is ISCO-coded, and Directive 2005/36/EC only covers
#    regulated professions. Caveat: 79 professions, EU/EEA-origin decisions only.
decisions = json.loads(
    (PACK / "eu-regulated-professions" / "regprof-decisions-germany.json").read_text(encoding="utf-8")
)
regulated = sorted({d["Isco code 1"] for d in decisions if d.get("Isco code 1")})

OUT.parent.mkdir(parents=True, exist_ok=True)
OUT.write_text(
    json.dumps(
        {
            "iscoToKldb": {k: sorted(v) for k, v in isco_to_kldb.items()},
            "kldbShortageScore": kldb_score,
            "regulatedIsco": regulated,
        }
    ),
    encoding="utf-8",
)
print(f"wrote {OUT}")
print(f"  isco groups:      {len(isco_to_kldb)}")
print(f"  scored kldb-4:    {len(kldb_score)}")
print(f"  regulated isco-4: {len(regulated)}")
